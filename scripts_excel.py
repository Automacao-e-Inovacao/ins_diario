import numpy as np
from openpyxl import load_workbook

def Manipulacao(Caminho_da_extracao_ccsupervision, cabecalho, piloto, data):
    lista_indices_dos_valores = []
    matriz_principal = []
    lista_indices_valores_inteiros = []
    wb = load_workbook(Caminho_da_extracao_ccsupervision)
    ws = wb.active
    matriz = np.asarray([[cell.value for cell in row] for row in ws.iter_rows()])
    
    for contagem_matriz, linha in enumerate(matriz):
        for contagem_valor_cabecalho, valor in enumerate(linha):
            if valor in cabecalho:
                lista_indices_dos_valores.append(contagem_valor_cabecalho)
                if valor == 'ChOf' or valor == 'ChA' or valor == 'TMA':
                    lista_indices_valores_inteiros.append(contagem_valor_cabecalho)
        if len(lista_indices_dos_valores)>0:
            if len(lista_indices_dos_valores) == len(cabecalho):
                for indice_de_valor in lista_indices_dos_valores:
                    if indice_de_valor in lista_indices_valores_inteiros:
                        try: matriz_principal.append(int(matriz[contagem_matriz+1][indice_de_valor]))
                        except: matriz_principal.append(None)
                    else:
                        matriz_principal.append(float(matriz[contagem_matriz+1][indice_de_valor]))
                matriz_principal.append(str(piloto))
                matriz_principal.append('{:04d}-{:02d}-{:02d}'.format(data.year, data.month, data.day))
                break
            lista_indices_dos_valores = []
    return (matriz_principal, matriz)

def Manipulacao_2(cabecalho, piloto, data, matriz):
    lista_indices_dos_valores = []
    matriz_principal = []
    lista_indices_valores_time = []
    lista_indices_valores_float = []
    lista_indices_valores_string = []
    lista_indices_valores_inteiro = []
    matriz_prov = []
    
    for contagem_matriz, linha in enumerate(matriz):
        
        for contagem_valor_cabecalho, valor in enumerate(linha):
            if valor in cabecalho:
                lista_indices_dos_valores.append(contagem_valor_cabecalho)
                if valor == 'Perini' or valor == 'Perfim':
                    lista_indices_valores_time.append(contagem_valor_cabecalho)
                elif valor == 'INS' or valor == 'IAb' or valor == 'ICO' or valor == 'Limite' or valor == 'Diferença' or valor == 'TME' or valor == 'TMA':
                    lista_indices_valores_float.append(contagem_valor_cabecalho)
                elif valor == 'Tipicidade':
                    lista_indices_valores_string.append(contagem_valor_cabecalho)
                else:
                    lista_indices_valores_inteiro.append(contagem_valor_cabecalho)
                    
        if len(lista_indices_dos_valores) == len(cabecalho):
            break
        else:
            lista_indices_dos_valores = []
        
    for linha_1 in matriz[contagem_matriz + 1:]:
        matriz_prov = []
        for indice_de_valor in lista_indices_dos_valores:
            if indice_de_valor in lista_indices_valores_time:
                matriz_prov.append(linha_1[indice_de_valor])
            elif indice_de_valor in lista_indices_valores_float:
                try: matriz_prov.append(float(linha_1[indice_de_valor]))
                except: matriz_prov.append(None)
            elif indice_de_valor in lista_indices_valores_string:
                if linha_1[indice_de_valor] != '':
                    matriz_prov.append(str(linha_1[indice_de_valor]))
            elif indice_de_valor in lista_indices_valores_inteiro:
                try: matriz_prov.append(int(linha_1[indice_de_valor]))
                except: matriz_prov.append(None)
        if not 'Típico' in matriz_prov and not 'Atípico' in matriz_prov:
            break
        matriz_prov.append('{:04d}-{:02d}-{:02d}'.format(data.year, data.month, data.day))
        matriz_prov.append(piloto)
        matriz_principal.append(matriz_prov)
    return matriz_principal